#!/usr/bin/env python3
import json
import sys
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl is required: python3 -m pip install openpyxl") from exc


def clean(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def compact_row(row):
    values = [clean(value) for value in row]
    while values and values[-1] in (None, ""):
        values.pop()
    return values


def classify_row(values):
    nonempty = [value for value in values if value not in (None, "")]
    if not nonempty:
        return "empty"
    first = str(nonempty[0]).strip().lower()
    if first in {"type", "level", "position"}:
        return "metadata"
    if any(str(value).strip().lower() in {"question", "answers", "correct", "explanation", "score", "results"} for value in nonempty):
        return "header"
    if len(nonempty) <= 3 and len(values) <= 4:
        return "key_value"
    return "data"


def inspect_sheet(ws, max_rows=120, max_cols=40):
    rows = []
    row_kinds = Counter()
    column_nonempty = Counter()

    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, max_rows), max_col=min(ws.max_column or 0, max_cols), values_only=True),
        1,
    ):
        values = compact_row(row)
        kind = classify_row(values)
        row_kinds[kind] += 1
        for column_index, value in enumerate(values, 1):
            if value not in (None, ""):
                column_nonempty[column_index] += 1
        if kind != "empty" and len(rows) < 40:
            rows.append({"row": row_index, "kind": kind, "values": values})

    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": [str(cell_range) for cell_range in ws.merged_cells.ranges],
        "row_kinds": dict(row_kinds),
        "column_nonempty": dict(sorted(column_nonempty.items())),
        "sample_rows": rows,
    }


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: inspect_workbook.py <workbook.xlsx>")

    path = Path(sys.argv[1])
    workbook = load_workbook(path, data_only=True, read_only=False)
    result = {
        "workbook": str(path),
        "sheets": [inspect_sheet(sheet) for sheet in workbook.worksheets],
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
